import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads"


def read_csv(name: str):
    with (DOWNLOADS / name).open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def normalize(value: str):
    value = unicodedata.normalize("NFKD", value.casefold()).encode("ascii", "ignore").decode()
    value = value.replace("&", " dan ")
    value = re.sub(r"\b(kepala dan anggota keluarga|kepala keluarga|anggota keluarga|responden)\b", " ", value)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def role_name(subject: str):
    if subject == "anggota keluarga":
        return "MEMBER"
    if subject == "kepala keluarga":
        return "HEAD"
    return "ALL"


def frequency_name(value: str):
    return {
        "Insidentil": "INCIDENTAL",
        "6 bulan": "SIX_MONTHS",
        "Tahunan": "ANNUAL",
        "Tidak Berubah": "IMMUTABLE",
    }.get(value)


mapping = json.loads((ROOT / "config" / "indikator-mapping.json").read_text(encoding="utf-8"))["kolom"]
merge_rows = read_csv("merge_pertanyaan.csv")
answer_rows = read_csv("jawabantertutup.csv")
question_rows = read_csv("questions(logic_jawaban).csv")

datatype_workbook = load_workbook(DOWNLOADS / "ajaib_datatype.xlsx", data_only=True)
datatype_sheet = datatype_workbook["ajaib_datatype"]
datatype_status = {}
for table, column, _data_type, status in datatype_sheet.iter_rows(min_row=2, values_only=True):
    if str(table or "").strip().casefold() == "ajaib" and column:
        datatype_status[str(column).strip()] = str(status or "").strip().casefold() or None

NON_EDITABLE_DATATYPE_STATUSES = {"old", "changed", "temporer", "perhitungan sistem"}

workbook = load_workbook(DOWNLOADS / "Parameter_DDP_Estimasi_Waktu_Updating (1).xlsx", data_only=True)
sheet = workbook["Semua Parameter"]
excel_rows = []
for row in sheet.iter_rows(min_row=4, values_only=True):
    if row[3] and frequency_name(str(row[5] or "").strip()):
        excel_rows.append(
            {
                "label": str(row[3]).strip(),
                "normalized": normalize(str(row[3])),
                "frequency": frequency_name(str(row[5]).strip()),
            }
        )

answers = defaultdict(list)
for row in answer_rows:
    if row["status"].strip() != "0":
        continue
    value = row["value"].strip()
    if value and value not in answers[row["pertanyaan_id"]]:
        answers[row["pertanyaan_id"]].append(value)

merge_by_id = {row["id"]: row for row in merge_rows}
questions_by_id = {row["id"]: row for row in question_rows}
merge_by_header = defaultdict(list)
for row in merge_rows:
    header = row["header"].strip()
    if header:
        merge_by_header[header].append(row)

EXPLICIT_LABELS = {
    "kode_deskel": "Kode Desa",
    "kode_bangunan": "Kode Bangunan",
    "deskel": "Desa/Kelurahan",
    "dusun": "Dusun/Kampung/ Dukuh",
    "rw": "RW",
    "rt": "RT",
    "responden": "Nama Responden",
    "kesediaan": "kesediaan sebagai responden",
    "nkk": "Nomor KK",
    "nama": "Nama Kepala Keluarga",
    "nik": "NIK Kepala Keluarga",
    "no_hp": "Nomor HP Responden",
    "dead_jml": "Jumlah Anggota Keluarga yang Meninggal",
    "thn_datang": "Tahun Datang ke Desa",
    "menetap": "Tinggal Menetap",
    "dinamika": "Dinamika Kehidupan",
    "tgl_lahir": "Tanggal Lahir Kepala Keluarga",
    "punya_aktalahir": "Kepemilikan Akta Lahir Kepala Keluarga dan Anggota Keluarga",
    "kendaraan_jml": "Jumlah Kendaraan Bermotor Dalam Rumah",
    "anak_asi": "ASI Eklusif",
    "rp_listrik": "Biaya Listrik Per bulan (Rp)",
    "rumah_kamar": "Jumah Kamar Tidur",
}


def field_frequency(field: str, rows):
    explicit = EXPLICIT_LABELS.get(field)
    if explicit:
        target = normalize(explicit)
        exact = next((row for row in excel_rows if row["normalized"] == target), None)
        if exact:
            return exact["frequency"], exact["label"], 1.0

    candidates = [normalize(row["value"]) for row in rows if row["value"].strip()]
    if not candidates:
        candidates = [normalize(field.replace("_", " "))]
    best = None
    for candidate in candidates:
        for excel in excel_rows:
            score = 1.0 if candidate == excel["normalized"] else SequenceMatcher(None, candidate, excel["normalized"]).ratio()
            if best is None or score > best[0]:
                best = (score, excel)
    if best and best[0] >= 0.90:
        return best[1]["frequency"], best[1]["label"], round(best[0], 3)
    return None, None, round(best[0], 3) if best else 0


fields = {}
unmapped_frequency = []
for field in mapping:
    rows = merge_by_header.get(field, [])
    frequency, source_label, confidence = field_frequency(field, rows)
    if not frequency:
        unmapped_frequency.append(field)
    variants = {}
    for row in rows:
        role = role_name(row["subjek"].strip())
        question = questions_by_id.get(row["id"])
        variant = variants.setdefault(role, {})
        variant["active"] = bool(variant.get("active")) or row["status"].strip() == "0"
        if answers[row["id"]]:
            variant["options"] = answers[row["id"]]
        if question and question["status"].strip() == "0":
            variant["inputType"] = question["question_type"].strip() or None
            description = " ".join(question["description"].split())
            if description:
                variant["help"] = description
            try:
                conditions = json.loads(question["conditions"] or "{}")
            except json.JSONDecodeError:
                conditions = {}
            show_if = conditions.get("show_if") if isinstance(conditions, dict) else None
            if isinstance(show_if, dict):
                source = merge_by_id.get(str(show_if.get("question_id")), {})
                source_field = source.get("header", "").strip()
                if source_field:
                    values = show_if.get("value", [])
                    variant["condition"] = {
                        "field": source_field,
                        "values": values if isinstance(values, list) else [values],
                    }
    status = datatype_status.get(field)
    datatype_editable = status not in NON_EDITABLE_DATATYPE_STATUSES
    fields[field] = {
        "frequency": frequency,
        "frequencySource": source_label,
        "frequencyConfidence": confidence,
        "datatypeStatus": status,
        "editable": datatype_editable and any(bool(variant.get("active")) for variant in variants.values()),
        "variants": variants,
    }

event_labels = {
    "deathCount": "Jumlah Anggota Keluarga yang Meninggal",
    "deathName": "Nama anggota keluarga yang meninggal",
    "deathNik": "NIK Anggota keluarga yang meninggal",
    "deathBirthDate": "Tanggal Lahir anggota keluarga yang meninggal",
    "deathDate": "Tanggal kematian anggota keluarga yang meninggal",
    "deathCause": "Penyebab Kematian anggota keluarga yang meninggal",
    "deathCertificate": "Kepemilikan Akta Kematian anggota keluarga yang meninggal",
}
events = {}
for key, label in event_labels.items():
    normalized = normalize(label)
    excel = next((row for row in excel_rows if row["normalized"] == normalized), None)
    merge = next((row for row in merge_rows if normalize(row["value"]) == normalized), None)
    events[key] = {
        "label": label,
        "frequency": excel["frequency"] if excel else "INCIDENTAL",
        "options": answers[merge["id"]] if merge else [],
    }

output = {
    "_meta": {
        "sources": [
            "Parameter_DDP_Estimasi_Waktu_Updating (1).xlsx",
            "jawabantertutup.csv",
            "merge_pertanyaan.csv",
            "questions(logic_jawaban).csv",
            "ajaib_datatype.xlsx",
        ],
        "tooltipPolicy": "QUESTION_LOGIC_ONLY",
        "unmappedFrequency": unmapped_frequency,
    },
    "fields": fields,
    "events": events,
    "building": {
        "kode": {"frequency": "ANNUAL"},
        "jenis": {"frequency": "ANNUAL"},
        "kategori": {"frequency": "ANNUAL"},
        "keterangan": {"frequency": "ANNUAL"},
        "fotoUrl": {"frequency": "ANNUAL"},
        "dusun": {"frequency": "ANNUAL"},
        "rw": {"frequency": "ANNUAL"},
        "rt": {"frequency": "ANNUAL"},
    },
}

(ROOT / "config" / "updating-metadata.json").write_text(
    json.dumps(output, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
)
print(f"Wrote metadata for {len(fields)} fields; {len(unmapped_frequency)} frequencies remain unspecified.")
